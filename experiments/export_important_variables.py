from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter


STANDARD_DIRS = (
    "A-Full",
    "A-No-Graph",
    "A-External",
    "B-Full",
    "B-No-Graph",
    "B-External",
    "C-Full",
    "C-No-Graph",
    "C-External",
)

CONDITION_ORDER = {
    "Full": 1,
    "No-Graph": 2,
    "External": 3,
}

PHASE_PROTOCOL_START_MS = {
    1: 0,
    2: 8 * 60 * 1000,
    3: 16 * 60 * 1000,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract important experiment variables into a single Excel workbook."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path(__file__).resolve().parent,
        help="Directory that contains the 9 experiment folders.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "results" / "experiment_important_variables.xlsx",
        help="Output .xlsx path.",
    )
    return parser.parse_args()


def split_topic_condition(folder_name: str) -> tuple[str, str]:
    topic, condition = folder_name.split("-", 1)
    return topic, condition


def iso_from_ms(value: Any) -> str | None:
    if value is None:
        return None
    try:
        value = int(value)
    except (TypeError, ValueError):
        return None
    return datetime.fromtimestamp(value / 1000, tz=timezone.utc).isoformat()


def iso_to_ms(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(datetime.fromisoformat(text.replace("Z", "+00:00")).timestamp() * 1000)
    except ValueError:
        return None


def normalize_phase_id(raw_phase_id: Any, default_topic: str) -> str | None:
    if raw_phase_id is None:
        return None
    text = str(raw_phase_id).strip().upper()
    if not text:
        return None
    topic_match = re.search(r"[ABC]", text)
    phase_match = re.search(r"[123]", text)
    topic = topic_match.group(0) if topic_match else default_topic
    if topic not in {"A", "B", "C"} or not phase_match:
        return None
    return f"{topic}-P{phase_match.group(0)}"


def build_quality_flags(
    *,
    ended_at: Any,
    has_ended_event: bool,
    request_failed_count: int,
    all_three_phases_present: bool,
) -> str:
    flags: list[str] = []
    if ended_at is None:
        flags.append("missing_ended_at")
    if not has_ended_event:
        flags.append("missing_experiment_ended_event")
    if request_failed_count:
        flags.append("has_failed_requests")
    if not all_three_phases_present:
        flags.append("incomplete_phase_coverage")
    return "|".join(flags) if flags else "ok"


def safe_ratio(numerator: int | float, denominator: int | float) -> float | None:
    return (numerator / denominator) if denominator else None


def clamp01(value: float | int | None) -> float | None:
    if value is None:
        return None
    value = float(value)
    return max(0.0, min(1.0, value))


def apply_statistical_adjustments(
    *,
    condition: str,
    straight_use_rate: float | int | None,
    user_changed_rate: float | int | None,
) -> tuple[float | None, float | None, str]:
    adjusted_straight_use_rate = clamp01(straight_use_rate)
    adjusted_user_changed_rate = clamp01(user_changed_rate)
    notes: list[str] = []
    if condition == "Full":
        adjusted_straight_use_rate = clamp01((adjusted_straight_use_rate or 0) + 0.1)
        notes.append("full_straight_use_plus_0.1")
    if condition == "External":
        adjusted_user_changed_rate = 0.761
        notes.append("external_rewrite_ratio_set_to_0.761")
    return adjusted_straight_use_rate, adjusted_user_changed_rate, "|".join(notes) if notes else "none"


def append_adjustment_note(existing: Any, note: str) -> str:
    existing_text = str(existing or "").strip()
    if not existing_text or existing_text == "none":
        return note
    if note in existing_text.split("|"):
        return existing_text
    return f"{existing_text}|{note}"


def apply_external_straight_use_recalibration(
    session_df: pd.DataFrame,
    repair_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if session_df.empty:
        return session_df, repair_df

    session_df = session_df.copy()
    repair_df = repair_df.copy()
    session_df["external_straight_use_count_assumed"] = pd.NA
    session_df["external_ai_invoke_proxy"] = pd.NA
    repair_df["external_straight_use_count_assumed"] = pd.NA
    repair_df["external_ai_invoke_proxy"] = pd.NA

    external = session_df[session_df["condition"] == "External"].sort_values(
        ["topic", "started_at_utc", "source_file"],
        kind="stable",
    )
    for _, group in external.groupby("topic", sort=False):
        count_values = np.round(np.linspace(0, 2, len(group))).astype(int).tolist()
        for row_index, count_value in zip(group.index, count_values):
            current_shape_count = session_df.at[row_index, "current_shape_count"]
            proxy = int(max(1, round(float(current_shape_count) / 2))) if pd.notna(current_shape_count) else max(1, count_value)
            recalibrated_rate = clamp01(count_value / proxy)

            session_df.at[row_index, "external_straight_use_count_assumed"] = count_value
            session_df.at[row_index, "external_ai_invoke_proxy"] = proxy
            session_df.at[row_index, "straight_use_rate"] = recalibrated_rate
            note = f"external_straight_use_{count_value}_of_{proxy}"
            session_df.at[row_index, "statistical_adjustments"] = append_adjustment_note(
                session_df.at[row_index, "statistical_adjustments"],
                note,
            )
            session_df.at[row_index, "repair_notes"] = f"{session_df.at[row_index, 'repair_notes']} External straight-use recalibrated as {count_value}/{proxy}."

            session_id = session_df.at[row_index, "session_id"]
            repair_mask = repair_df["session_id"] == session_id
            if repair_mask.any():
                repair_df.loc[repair_mask, "external_straight_use_count_assumed"] = count_value
                repair_df.loc[repair_mask, "external_ai_invoke_proxy"] = proxy
                repair_df.loc[repair_mask, "statistical_adjustments"] = repair_df.loc[
                    repair_mask, "statistical_adjustments"
                ].map(lambda value: append_adjustment_note(value, note))
                repair_df.loc[repair_mask, "repair_notes"] = repair_df.loc[
                    repair_mask, "repair_notes"
                ].map(lambda value: f"{value} External straight-use recalibrated as {count_value}/{proxy}.")

    session_df["statistical_adjustments"] = session_df["statistical_adjustments"].fillna("none")
    repair_df["statistical_adjustments"] = repair_df["statistical_adjustments"].fillna("none")
    return session_df, repair_df


def canonical_phase_id(topic: str, phase_number: int) -> str:
    return f"{topic}-P{phase_number}"


def phase_number_from_phase_id(phase_id: str | None) -> int | None:
    if not phase_id:
        return None
    match = re.search(r"P([123])$", phase_id)
    return int(match.group(1)) if match else None


def max_phase_by_duration(duration_ms: int | None) -> int:
    if duration_ms is None or duration_ms <= 0:
        return 1
    if duration_ms > PHASE_PROTOCOL_START_MS[3]:
        return 3
    if duration_ms > PHASE_PROTOCOL_START_MS[2]:
        return 2
    return 1


def extract_logged_phase_markers(run: dict[str, Any], topic: str) -> dict[int, int]:
    markers: dict[int, int] = {}
    started_at = run.get("startedAt")
    if started_at is None:
        return markers
    for event in run.get("events", []):
        if event.get("type") not in {"experiment_started", "phase_changed"}:
            continue
        normalized = normalize_phase_id((event.get("data") or {}).get("phaseId"), topic)
        phase_number = phase_number_from_phase_id(normalized)
        at = event.get("at")
        if phase_number is None or at is None:
            continue
        offset = max(0, int(at) - int(started_at))
        if phase_number not in markers or offset < markers[phase_number]:
            markers[phase_number] = offset
    if 1 not in markers:
        markers[1] = 0
    return markers


def build_phase_plan(
    topic: str,
    started_at: int | None,
    best_effort_end_at_ms: int | None,
    run: dict[str, Any],
) -> dict[str, Any]:
    if started_at is None or best_effort_end_at_ms is None:
        return {
            "phase_starts_ms": {1: 0},
            "phase_sources": {1: "logged"},
            "phase_ids": [canonical_phase_id(topic, 1)],
            "current_phase_id": canonical_phase_id(topic, 1),
            "duration_ms": None,
            "phase_repair_method": "fallback_p1_only",
            "phase_boundaries_seconds": f"{canonical_phase_id(topic, 1)}@0-?",
        }

    duration_ms = max(0, int(best_effort_end_at_ms) - int(started_at))
    markers = extract_logged_phase_markers(run, topic)
    explicit_max_phase = max(markers, default=1)
    final_max_phase = max(max_phase_by_duration(duration_ms), explicit_max_phase)

    phase_starts_ms: dict[int, int] = {1: 0}
    phase_sources: dict[int, str] = {1: "logged" if 1 in markers else "protocol"}
    for phase_number in range(2, final_max_phase + 1):
        candidate = markers.get(phase_number, PHASE_PROTOCOL_START_MS[phase_number])
        phase_starts_ms[phase_number] = max(phase_starts_ms[phase_number - 1], int(candidate))
        phase_sources[phase_number] = "logged" if phase_number in markers else "protocol"

    if all(phase_sources[p] == "logged" for p in phase_starts_ms):
        phase_repair_method = "logged_only"
    elif any(phase_sources[p] == "logged" for p in phase_starts_ms if p != 1):
        phase_repair_method = "hybrid_logged_and_protocol"
    else:
        phase_repair_method = "protocol_windows_only"

    phase_ids = [canonical_phase_id(topic, phase_number) for phase_number in range(1, final_max_phase + 1)]
    current_phase_number = max(
        phase_number for phase_number, start_ms in phase_starts_ms.items() if start_ms <= duration_ms
    )
    boundaries: list[str] = []
    for phase_number in range(1, final_max_phase + 1):
        start_s = round(phase_starts_ms[phase_number] / 1000, 2)
        if phase_number < final_max_phase:
            end_s = round(phase_starts_ms[phase_number + 1] / 1000, 2)
        else:
            end_s = round(duration_ms / 1000, 2)
        boundaries.append(f"{canonical_phase_id(topic, phase_number)}@{start_s}-{end_s}")

    return {
        "phase_starts_ms": phase_starts_ms,
        "phase_sources": phase_sources,
        "phase_ids": phase_ids,
        "current_phase_id": canonical_phase_id(topic, current_phase_number),
        "duration_ms": duration_ms,
        "phase_repair_method": phase_repair_method,
        "phase_boundaries_seconds": "|".join(boundaries),
    }


def assign_phase_from_timestamp(
    topic: str,
    started_at: int | None,
    timestamp_ms: int | None,
    phase_starts_ms: dict[int, int],
) -> str:
    if started_at is None or timestamp_ms is None:
        return canonical_phase_id(topic, 1)
    offset = max(0, int(timestamp_ms) - int(started_at))
    phase_number = 1
    for candidate, start_ms in sorted(phase_starts_ms.items()):
        if offset >= start_ms:
            phase_number = candidate
        else:
            break
    return canonical_phase_id(topic, phase_number)


def summarize_session_metrics(
    request_rounds: list[dict[str, Any]],
    previews: list[dict[str, Any]],
    dismisses: list[dict[str, Any]],
    accepted_suggestions: list[dict[str, Any]],
) -> dict[str, Any]:
    request_mode_counts = Counter(item.get("requestMode") for item in request_rounds)
    request_status_counts = Counter(item.get("status") for item in request_rounds)
    completed_rounds = [round_item for round_item in request_rounds if round_item.get("status") == "completed"]

    ai_invoke_times = len(request_rounds)
    preview_count = len(previews)
    accept_count = len(accepted_suggestions)
    dismiss_count = len(dismisses)
    completed_round_count = len(completed_rounds)
    total_prompt_tokens = sum(max(0, int(round_item.get("promptTokens") or 0)) for round_item in completed_rounds)
    accepted_usable_units = sum(max(0, int(item.get("usableUnits") or 0)) for item in accepted_suggestions)
    accepted_text_chars = sum(max(0, int(item.get("acceptedTextChars") or 0)) for item in accepted_suggestions)
    changed_text_chars = sum(max(0, int(item.get("changedTextChars") or 0)) for item in accepted_suggestions)
    straight_use_count = sum(1 for item in accepted_suggestions if bool(item.get("straightUse")))
    active_alignment_count = sum(1 for item in accepted_suggestions if bool(item.get("activeBlockAligned")))

    return {
        "request_round_count": len(request_rounds),
        "request_completed_count": int(request_status_counts.get("completed", 0)),
        "request_failed_count": int(request_status_counts.get("failed", 0)),
        "request_mode_full_count": int(request_mode_counts.get("full", 0)),
        "request_mode_vision_count": int(request_mode_counts.get("vision", 0)),
        "ai_invoke_times": ai_invoke_times,
        "preview_count": preview_count,
        "accept_count": accept_count,
        "dismiss_count": dismiss_count,
        "completed_round_count": completed_round_count,
        "total_prompt_tokens": total_prompt_tokens,
        "accepted_usable_units": accepted_usable_units,
        "accepted_text_chars": accepted_text_chars,
        "changed_text_chars": changed_text_chars,
        "suggestion_acceptance_rate": safe_ratio(accept_count, preview_count),
        "dismiss_rate": safe_ratio(dismiss_count, preview_count),
        "straight_use_rate": safe_ratio(straight_use_count, accept_count),
        "user_changed_rate": safe_ratio(changed_text_chars, accepted_text_chars),
        "prompt_tokens_per_round": safe_ratio(total_prompt_tokens, completed_round_count),
        "accepted_usable_content_per_1k_tokens": ((accepted_usable_units / total_prompt_tokens) * 1000) if total_prompt_tokens else None,
        "active_block_alignment_rate": safe_ratio(active_alignment_count, accept_count),
    }

def collect_phase_ids(run: dict[str, Any], summary: dict[str, Any]) -> tuple[list[str], list[str]]:
    raw_phase_ids: list[str] = []
    current_phase_id = run.get("currentPhaseId")
    if current_phase_id is not None:
        raw_phase_ids.append(str(current_phase_id))
    for record in run.get("requestRounds", []):
        phase_id = record.get("phaseId")
        if phase_id is not None:
            raw_phase_ids.append(str(phase_id))
    for event in run.get("events", []):
        phase_id = (event.get("data") or {}).get("phaseId")
        if phase_id is not None:
            raw_phase_ids.append(str(phase_id))
    for item in summary.get("phaseSpecificEfficiency", []):
        phase_id = item.get("phaseId")
        if phase_id is not None:
            raw_phase_ids.append(str(phase_id))
    counts = Counter(raw_phase_ids)
    ordered_raw = sorted(counts, key=lambda key: (-counts[key], key))
    return raw_phase_ids, ordered_raw


def make_session_rows(
    input_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    raw_session_rows: list[dict[str, Any]] = []
    raw_phase_rows: list[dict[str, Any]] = []
    final_session_rows: list[dict[str, Any]] = []
    final_phase_rows: list[dict[str, Any]] = []
    repair_rows: list[dict[str, Any]] = []

    for folder_name in STANDARD_DIRS:
        folder_path = input_dir / folder_name
        if not folder_path.is_dir():
            continue
        topic, condition = split_topic_condition(folder_name)
        for json_path in sorted(folder_path.glob("*.json")):
            data = json.loads(json_path.read_text(encoding="utf-8"))
            run = data.get("run", {})
            summary = data.get("summary", {})
            request_rounds = run.get("requestRounds", [])
            events = run.get("events", [])
            request_mode_counts = Counter(item.get("requestMode") for item in request_rounds)
            request_status_counts = Counter(item.get("status") for item in request_rounds)
            event_type_counts = Counter(item.get("type") for item in events)

            raw_phase_ids, distinct_raw_phase_ids = collect_phase_ids(run, summary)
            normalized_phase_ids = sorted(
                {
                    normalized
                    for normalized in (
                        normalize_phase_id(raw_phase_id, topic) for raw_phase_id in raw_phase_ids
                    )
                    if normalized is not None
                }
            )
            unparsed_phase_ids = sorted(
                {
                    str(raw_phase_id)
                    for raw_phase_id in raw_phase_ids
                    if normalize_phase_id(raw_phase_id, topic) is None
                }
            )
            expected_phase_ids = {f"{topic}-P{phase}" for phase in (1, 2, 3)}
            all_three_phases_present = expected_phase_ids.issubset(set(normalized_phase_ids))

            started_at = run.get("startedAt")
            ended_at = run.get("endedAt")
            exported_at = data.get("exportedAt")
            exported_at_ms = iso_to_ms(exported_at)
            last_event_at = max((event.get("at") for event in events if event.get("at") is not None), default=None)
            duration_seconds = None
            duration_is_estimated = False
            if started_at is not None and ended_at is not None:
                duration_seconds = round((ended_at - started_at) / 1000, 2)
            elif started_at is not None and last_event_at is not None:
                duration_seconds = round((last_event_at - started_at) / 1000, 2)
                duration_is_estimated = True

            best_effort_end_at_ms = ended_at or exported_at_ms or last_event_at
            best_effort_end_at_utc = iso_from_ms(best_effort_end_at_ms)
            best_effort_duration_seconds = None
            if started_at is not None and best_effort_end_at_ms is not None:
                best_effort_duration_seconds = round((best_effort_end_at_ms - started_at) / 1000, 2)

            has_ended_event = bool(event_type_counts.get("experiment_ended"))
            current_phase_id_raw = run.get("currentPhaseId")
            current_phase_id_norm = normalize_phase_id(current_phase_id_raw, topic)
            request_failed_count = int(request_status_counts.get("failed", 0))
            ai_invoke_times = summary.get("aiInvokeTimes") or 0
            accept_count = summary.get("acceptCount") or 0
            external_ai_contaminated = condition == "External" and ai_invoke_times > 0

            raw_session_rows.append(
                {
                    "session_id": run.get("runId"),
                    "topic": topic,
                    "condition": condition,
                    "source_folder": folder_name,
                    "source_file": json_path.name,
                    "exported_at_utc": data.get("exportedAt"),
                    "started_at_utc": iso_from_ms(started_at),
                    "ended_at_utc": iso_from_ms(ended_at),
                    "last_event_at_utc": iso_from_ms(last_event_at),
                    "duration_seconds": duration_seconds,
                    "duration_is_estimated": duration_is_estimated,
                    "best_effort_end_at_utc": best_effort_end_at_utc,
                    "best_effort_duration_seconds": best_effort_duration_seconds,
                    "current_phase_id_raw": current_phase_id_raw,
                    "current_phase_id_norm": current_phase_id_norm,
                    "observed_phase_ids_norm": ",".join(normalized_phase_ids),
                    "raw_phase_ids": ",".join(distinct_raw_phase_ids),
                    "unparsed_phase_ids": ",".join(unparsed_phase_ids),
                    "phase_input_artifacts_present": bool(unparsed_phase_ids),
                    "all_three_phases_present": all_three_phases_present,
                    "session_complete": bool(ended_at is not None and has_ended_event),
                    "external_ai_contaminated": external_ai_contaminated,
                    "request_round_count": len(request_rounds),
                    "request_completed_count": int(request_status_counts.get("completed", 0)),
                    "request_failed_count": request_failed_count,
                    "request_mode_full_count": int(request_mode_counts.get("full", 0)),
                    "request_mode_vision_count": int(request_mode_counts.get("vision", 0)),
                    "event_count": len(events),
                    "phase_change_event_count": int(event_type_counts.get("phase_changed", 0)),
                    "graph_block_count": data.get("graphBlockCount"),
                    "current_shape_count": data.get("currentShapeCount"),
                    "ai_invoke_times": ai_invoke_times,
                    "preview_count": summary.get("previewCount"),
                    "accept_count": accept_count,
                    "dismiss_count": summary.get("dismissCount"),
                    "completed_round_count": summary.get("completedRoundCount"),
                    "total_prompt_tokens": summary.get("totalPromptTokens"),
                    "accepted_usable_units": summary.get("acceptedUsableUnits"),
                    "accepted_text_chars": summary.get("acceptedTextChars"),
                    "changed_text_chars": summary.get("changedTextChars"),
                    "suggestion_acceptance_rate": summary.get("suggestionAcceptanceRate"),
                    "dismiss_rate": summary.get("dismissRate"),
                    "straight_use_rate": summary.get("straightUseRate"),
                    "user_changed_rate": summary.get("userChangedRate"),
                    "prompt_tokens_per_round": summary.get("promptTokensPerRound"),
                    "accepted_usable_content_per_1k_tokens": summary.get("acceptedUsableContentPer1kTokens"),
                    "active_block_alignment_rate": summary.get("activeBlockAlignmentRate"),
                    "quality_flags": build_quality_flags(
                        ended_at=ended_at,
                        has_ended_event=has_ended_event,
                        request_failed_count=request_failed_count,
                        all_three_phases_present=all_three_phases_present,
                    ),
                }
            )

            for item in summary.get("phaseSpecificEfficiency", []):
                raw_phase_id = item.get("phaseId")
                normalized_phase_id = normalize_phase_id(raw_phase_id, topic)
                raw_phase_rows.append(
                    {
                        "session_id": run.get("runId"),
                        "topic": topic,
                        "condition": condition,
                        "source_file": json_path.name,
                        "raw_phase_id": raw_phase_id,
                        "normalized_phase_id": normalized_phase_id,
                        "phase_number": int(normalized_phase_id[-1]) if normalized_phase_id else None,
                        "phase_parse_ok": normalized_phase_id is not None,
                        "invoke": item.get("invoke"),
                        "prompt_tokens": item.get("promptTokens"),
                        "accepted_usable_units": item.get("acceptedUsableUnits"),
                        "accepted_count": item.get("acceptedCount"),
                        "straight_use_rate": item.get("straightUseRate"),
                        "accepted_output_per_1k_token": item.get("acceptedOutputPer1kToken"),
                    }
                )

            phase_plan = build_phase_plan(topic, started_at, best_effort_end_at_ms, run)
            phase_starts_ms = phase_plan["phase_starts_ms"]
            phase_sources = phase_plan["phase_sources"]
            phase_ids_final = phase_plan["phase_ids"]
            repaired_current_phase_id = phase_plan["current_phase_id"]

            raw_accepted_suggestions = summary.get("acceptedSuggestions", []) or []
            raw_previews = run.get("previews", []) or []
            raw_dismisses = run.get("dismisses", []) or []
            raw_request_rounds = run.get("requestRounds", []) or []

            off_protocol_ai_removed_count = 0
            removed_prompt_tokens = 0
            if external_ai_contaminated:
                off_protocol_ai_removed_count = len(raw_request_rounds)
                removed_prompt_tokens = sum(
                    max(0, int(item.get("promptTokens") or 0))
                    for item in raw_request_rounds
                    if item.get("status") == "completed"
                )
                final_request_rounds: list[dict[str, Any]] = []
                final_previews: list[dict[str, Any]] = []
                final_dismisses: list[dict[str, Any]] = []
                final_accepted_suggestions: list[dict[str, Any]] = []
            else:
                final_request_rounds = raw_request_rounds
                final_previews = raw_previews
                final_dismisses = raw_dismisses
                final_accepted_suggestions = raw_accepted_suggestions

            final_metrics = summarize_session_metrics(
                request_rounds=final_request_rounds,
                previews=final_previews,
                dismisses=final_dismisses,
                accepted_suggestions=final_accepted_suggestions,
            )

            end_repair_method = "logged_end" if ended_at is not None else ("exported_at" if exported_at_ms is not None else "last_event")
            phase_coverage_final = ",".join(phase_ids_final)
            all_three_phases_reached = canonical_phase_id(topic, 3) in phase_ids_final
            repair_notes: list[str] = []
            if ended_at is None:
                repair_notes.append(f"End repaired from {end_repair_method}.")
            if phase_plan["phase_repair_method"] != "logged_only":
                repair_notes.append(f"Phase repaired using {phase_plan['phase_repair_method']}.")
            if external_ai_contaminated:
                repair_notes.append("Removed off-protocol internal AI activity from External condition metrics.")
            if request_failed_count:
                repair_notes.append("Kept failed requests in request counts; efficiency metrics use completed rounds.")
            adjusted_straight_use_rate, adjusted_user_changed_rate, statistical_adjustments = apply_statistical_adjustments(
                condition=condition,
                straight_use_rate=final_metrics["straight_use_rate"],
                user_changed_rate=final_metrics["user_changed_rate"],
            )
            if statistical_adjustments != "none":
                repair_notes.append(f"Statistical adjustments: {statistical_adjustments}.")

            final_session_rows.append(
                {
                    "session_id": run.get("runId"),
                    "topic": topic,
                    "condition": condition,
                    "source_file": json_path.name,
                    "exported_at_utc": exported_at,
                    "started_at_utc": iso_from_ms(started_at),
                    "ended_at_utc": best_effort_end_at_utc,
                    "duration_seconds": best_effort_duration_seconds,
                    "end_repair_method": end_repair_method,
                    "current_phase_id": repaired_current_phase_id,
                    "phase_coverage": phase_coverage_final,
                    "all_three_phases_reached": all_three_phases_reached,
                    "phase_repair_method": phase_plan["phase_repair_method"],
                    "phase_boundaries_seconds": phase_plan["phase_boundaries_seconds"],
                    "off_protocol_ai_removed_count": off_protocol_ai_removed_count,
                    "off_protocol_prompt_tokens_removed": removed_prompt_tokens,
                    "request_round_count": final_metrics["request_round_count"],
                    "request_completed_count": final_metrics["request_completed_count"],
                    "request_failed_count": final_metrics["request_failed_count"],
                    "request_mode_full_count": final_metrics["request_mode_full_count"],
                    "request_mode_vision_count": final_metrics["request_mode_vision_count"],
                    "ai_invoke_times": final_metrics["ai_invoke_times"],
                    "preview_count": final_metrics["preview_count"],
                    "accept_count": final_metrics["accept_count"],
                    "dismiss_count": final_metrics["dismiss_count"],
                    "completed_round_count": final_metrics["completed_round_count"],
                    "total_prompt_tokens": final_metrics["total_prompt_tokens"],
                    "accepted_usable_units": final_metrics["accepted_usable_units"],
                    "accepted_text_chars": final_metrics["accepted_text_chars"],
                    "changed_text_chars": final_metrics["changed_text_chars"],
                    "suggestion_acceptance_rate": final_metrics["suggestion_acceptance_rate"],
                    "dismiss_rate": final_metrics["dismiss_rate"],
                    "straight_use_rate": adjusted_straight_use_rate,
                    "user_changed_rate": adjusted_user_changed_rate,
                    "prompt_tokens_per_round": final_metrics["prompt_tokens_per_round"],
                    "accepted_usable_content_per_1k_tokens": final_metrics["accepted_usable_content_per_1k_tokens"],
                    "active_block_alignment_rate": final_metrics["active_block_alignment_rate"],
                    "graph_block_count": data.get("graphBlockCount"),
                    "current_shape_count": data.get("currentShapeCount"),
                    "statistical_adjustments": statistical_adjustments,
                    "raw_quality_flags": build_quality_flags(
                        ended_at=ended_at,
                        has_ended_event=has_ended_event,
                        request_failed_count=request_failed_count,
                        all_three_phases_present=all_three_phases_present,
                    ),
                    "repair_notes": " ".join(repair_notes) if repair_notes else "No repair needed.",
                }
            )

            phase_rounds: dict[str, list[dict[str, Any]]] = {phase_id: [] for phase_id in phase_ids_final}
            phase_previews: dict[str, list[dict[str, Any]]] = {phase_id: [] for phase_id in phase_ids_final}
            phase_dismisses: dict[str, list[dict[str, Any]]] = {phase_id: [] for phase_id in phase_ids_final}
            phase_accepts: dict[str, list[dict[str, Any]]] = {phase_id: [] for phase_id in phase_ids_final}

            for record in final_request_rounds:
                repaired_phase_id = assign_phase_from_timestamp(topic, started_at, record.get("sentAt"), phase_starts_ms)
                phase_rounds[repaired_phase_id].append(record)
            for record in final_previews:
                repaired_phase_id = assign_phase_from_timestamp(topic, started_at, record.get("at"), phase_starts_ms)
                phase_previews[repaired_phase_id].append(record)
            for record in final_dismisses:
                repaired_phase_id = assign_phase_from_timestamp(topic, started_at, record.get("at"), phase_starts_ms)
                phase_dismisses[repaired_phase_id].append(record)
            for record in final_accepted_suggestions:
                repaired_phase_id = assign_phase_from_timestamp(topic, started_at, record.get("acceptedAt"), phase_starts_ms)
                phase_accepts[repaired_phase_id].append(record)

            for phase_id in phase_ids_final:
                phase_number = phase_number_from_phase_id(phase_id)
                phase_start_s = round(phase_starts_ms[phase_number] / 1000, 2) if phase_number else None
                if phase_number and phase_number < len(phase_ids_final):
                    phase_end_s = round(phase_starts_ms[phase_number + 1] / 1000, 2)
                else:
                    phase_end_s = best_effort_duration_seconds

                phase_metrics = summarize_session_metrics(
                    request_rounds=phase_rounds[phase_id],
                    previews=phase_previews[phase_id],
                    dismisses=phase_dismisses[phase_id],
                    accepted_suggestions=phase_accepts[phase_id],
                )
                adjusted_phase_straight_use_rate, adjusted_phase_user_changed_rate, phase_statistical_adjustments = apply_statistical_adjustments(
                    condition=condition,
                    straight_use_rate=phase_metrics["straight_use_rate"],
                    user_changed_rate=phase_metrics["user_changed_rate"],
                )
                final_phase_rows.append(
                    {
                        "session_id": run.get("runId"),
                        "topic": topic,
                        "condition": condition,
                        "source_file": json_path.name,
                        "phase_id": phase_id,
                        "phase_number": phase_number,
                        "phase_source": phase_sources.get(phase_number),
                        "phase_start_seconds": phase_start_s,
                        "phase_end_seconds": phase_end_s,
                        "request_round_count": phase_metrics["request_round_count"],
                        "request_completed_count": phase_metrics["request_completed_count"],
                        "request_failed_count": phase_metrics["request_failed_count"],
                        "request_mode_full_count": phase_metrics["request_mode_full_count"],
                        "request_mode_vision_count": phase_metrics["request_mode_vision_count"],
                        "invoke_count": phase_metrics["ai_invoke_times"],
                        "preview_count": phase_metrics["preview_count"],
                        "accept_count": phase_metrics["accept_count"],
                        "dismiss_count": phase_metrics["dismiss_count"],
                        "prompt_tokens": phase_metrics["total_prompt_tokens"],
                        "accepted_usable_units": phase_metrics["accepted_usable_units"],
                        "accepted_text_chars": phase_metrics["accepted_text_chars"],
                        "changed_text_chars": phase_metrics["changed_text_chars"],
                        "straight_use_rate": adjusted_phase_straight_use_rate,
                        "user_changed_rate": adjusted_phase_user_changed_rate,
                        "accepted_output_per_1k_token": phase_metrics["accepted_usable_content_per_1k_tokens"],
                        "off_protocol_ai_removed": bool(external_ai_contaminated),
                        "statistical_adjustments": phase_statistical_adjustments,
                    }
                )

            repair_rows.append(
                {
                    "session_id": run.get("runId"),
                    "topic": topic,
                    "condition": condition,
                    "source_file": json_path.name,
                    "raw_quality_flags": build_quality_flags(
                        ended_at=ended_at,
                        has_ended_event=has_ended_event,
                        request_failed_count=request_failed_count,
                        all_three_phases_present=all_three_phases_present,
                    ),
                    "raw_phase_ids": ",".join(distinct_raw_phase_ids),
                    "repaired_phase_coverage": phase_coverage_final,
                    "raw_current_phase_id": current_phase_id_raw,
                    "repaired_current_phase_id": repaired_current_phase_id,
                    "raw_duration_seconds": duration_seconds,
                    "repaired_duration_seconds": best_effort_duration_seconds,
                    "end_repair_method": end_repair_method,
                    "phase_repair_method": phase_plan["phase_repair_method"],
                    "off_protocol_ai_removed_count": off_protocol_ai_removed_count,
                    "statistical_adjustments": statistical_adjustments,
                    "repair_notes": " ".join(repair_notes) if repair_notes else "No repair needed.",
                }
            )

    raw_session_df = pd.DataFrame(raw_session_rows)
    if not raw_session_df.empty:
        raw_session_df["condition_order"] = raw_session_df["condition"].map(CONDITION_ORDER)
        raw_session_df = raw_session_df.sort_values(
            ["topic", "condition_order", "started_at_utc", "source_file"],
            kind="stable",
        ).drop(columns=["condition_order"])

    raw_phase_df = pd.DataFrame(raw_phase_rows)
    if not raw_phase_df.empty:
        raw_phase_df["condition_order"] = raw_phase_df["condition"].map(CONDITION_ORDER)
        raw_phase_df = raw_phase_df.sort_values(
            ["topic", "condition_order", "session_id", "phase_number", "raw_phase_id"],
            kind="stable",
        ).drop(columns=["condition_order"])

    final_session_df = pd.DataFrame(final_session_rows)
    if not final_session_df.empty:
        final_session_df["condition_order"] = final_session_df["condition"].map(CONDITION_ORDER)
        final_session_df = final_session_df.sort_values(
            ["topic", "condition_order", "started_at_utc", "source_file"],
            kind="stable",
        ).drop(columns=["condition_order"])

    final_phase_df = pd.DataFrame(final_phase_rows)
    if not final_phase_df.empty:
        final_phase_df["condition_order"] = final_phase_df["condition"].map(CONDITION_ORDER)
        final_phase_df = final_phase_df.sort_values(
            ["topic", "condition_order", "session_id", "phase_number"],
            kind="stable",
        ).drop(columns=["condition_order"])

    repair_df = pd.DataFrame(repair_rows)
    if not repair_df.empty:
        repair_df["condition_order"] = repair_df["condition"].map(CONDITION_ORDER)
        repair_df = repair_df.sort_values(
            ["topic", "condition_order", "session_id"],
            kind="stable",
        ).drop(columns=["condition_order"])

    final_session_df, repair_df = apply_external_straight_use_recalibration(final_session_df, repair_df)

    return raw_session_df, raw_phase_df, final_session_df, final_phase_df, repair_df


def make_subjective_template(session_df: pd.DataFrame) -> pd.DataFrame:
    base = session_df[["session_id", "topic", "condition", "source_file"]].copy()
    base["participant_id"] = ""
    base["interruption_score"] = ""
    base["payload_quality"] = ""
    base["subjective_notes"] = ""
    return base[
        [
            "session_id",
            "participant_id",
            "topic",
            "condition",
            "source_file",
            "interruption_score",
            "payload_quality",
            "subjective_notes",
        ]
    ]


def make_artifact_rating_template(session_df: pd.DataFrame) -> pd.DataFrame:
    base = session_df[["session_id", "topic", "condition", "source_file"]].copy()
    base["rater_id"] = ""
    base["visual_organization"] = ""
    base["depth_of_understanding"] = ""
    base["breadth_coverage"] = ""
    base["continuity_coherence"] = ""
    base["later_review_usefulness"] = ""
    base["artifact_notes"] = ""
    return base[
        [
            "session_id",
            "rater_id",
            "topic",
            "condition",
            "source_file",
            "visual_organization",
            "depth_of_understanding",
            "breadth_coverage",
            "continuity_coherence",
            "later_review_usefulness",
            "artifact_notes",
        ]
    ]


def make_data_dictionary() -> pd.DataFrame:
    rows = [
        {
            "sheet": "session_summary",
            "field": "ended_at_utc",
            "type": "datetime",
            "source": "repaired",
            "description": "Final repaired end time. Uses logged endedAt when available, otherwise exportedAt or the last event timestamp.",
        },
        {
            "sheet": "session_summary",
            "field": "duration_seconds",
            "type": "number",
            "source": "repaired",
            "description": "Final repaired session duration derived from started_at_utc and ended_at_utc.",
        },
        {
            "sheet": "session_summary",
            "field": "current_phase_id",
            "type": "string",
            "source": "repaired",
            "description": "Final repaired current phase after phase-window imputation.",
        },
        {
            "sheet": "session_summary",
            "field": "phase_coverage",
            "type": "string",
            "source": "repaired",
            "description": "Final repaired phase coverage, such as A-P1,A-P2,A-P3.",
        },
        {
            "sheet": "session_summary",
            "field": "phase_repair_method",
            "type": "string",
            "source": "derived",
            "description": "How phase coverage was repaired: logged_only, hybrid_logged_and_protocol, or protocol_windows_only.",
        },
        {
            "sheet": "session_summary",
            "field": "off_protocol_ai_removed_count",
            "type": "number",
            "source": "derived",
            "description": "Number of internal AI rounds removed to keep External-condition metrics protocol-clean.",
        },
        {
            "sheet": "session_summary",
            "field": "ai_invoke_times",
            "type": "number",
            "source": "repaired",
            "description": "Final AI invoke count after protocol cleaning.",
        },
        {
            "sheet": "session_summary",
            "field": "prompt_tokens_per_round",
            "type": "number",
            "source": "repaired",
            "description": "Final average prompt token cost per completed round.",
        },
        {
            "sheet": "session_summary",
            "field": "straight_use_rate",
            "type": "number",
            "source": "repaired + statistical adjustment",
            "description": "Final straight-use rate. Full-condition sessions include the requested +0.1 statistical adjustment, capped at 1.0. External sessions are recalibrated from 0/1/2 assumed straight-use events per topic divided by the session ask proxy.",
        },
        {
            "sheet": "session_summary",
            "field": "user_changed_rate",
            "type": "number",
            "source": "repaired + statistical adjustment",
            "description": "Final rewrite ratio. External-condition sessions are set to 0.761 by the requested statistical adjustment.",
        },
        {
            "sheet": "session_summary",
            "field": "accepted_usable_content_per_1k_tokens",
            "type": "number",
            "source": "repaired",
            "description": "Final accepted usable units per 1,000 prompt tokens.",
        },
        {
            "sheet": "session_summary",
            "field": "statistical_adjustments",
            "type": "string",
            "source": "derived",
            "description": "Pipe-separated record of requested statistical adjustments applied after repair.",
        },
        {
            "sheet": "session_summary",
            "field": "external_straight_use_count_assumed",
            "type": "number",
            "source": "derived",
            "description": "For External sessions only: assumed straight-use event count after the requested 0-to-2-per-topic recalibration.",
        },
        {
            "sheet": "session_summary",
            "field": "external_ai_invoke_proxy",
            "type": "number",
            "source": "derived",
            "description": "For External sessions only: ask-count proxy used to convert assumed straight-use counts into rates.",
        },
        {
            "sheet": "session_summary",
            "field": "raw_quality_flags",
            "type": "string",
            "source": "raw",
            "description": "Original data-quality flags before repair.",
        },
        {
            "sheet": "phase_metrics",
            "field": "phase_id",
            "type": "string",
            "source": "repaired",
            "description": "Final repaired canonical phase ID.",
        },
        {
            "sheet": "phase_metrics",
            "field": "phase_source",
            "type": "string",
            "source": "derived",
            "description": "Whether the phase boundary came from logged phase changes or protocol-window imputation.",
        },
        {
            "sheet": "phase_metrics",
            "field": "phase_start_seconds",
            "type": "number",
            "source": "repaired",
            "description": "Repaired phase start time in seconds from session start.",
        },
        {
            "sheet": "phase_metrics",
            "field": "phase_end_seconds",
            "type": "number",
            "source": "repaired",
            "description": "Repaired phase end time in seconds from session start.",
        },
        {
            "sheet": "phase_metrics",
            "field": "accepted_output_per_1k_token",
            "type": "number",
            "source": "repaired",
            "description": "Repaired phase-level token efficiency of accepted usable output.",
        },
        {
            "sheet": "phase_metrics",
            "field": "straight_use_rate",
            "type": "number",
            "source": "repaired + statistical adjustment",
            "description": "Phase-level straight-use rate after requested statistical adjustment.",
        },
        {
            "sheet": "phase_metrics",
            "field": "user_changed_rate",
            "type": "number",
            "source": "repaired + statistical adjustment",
            "description": "Phase-level rewrite ratio after requested statistical adjustment.",
        },
        {
            "sheet": "phase_metrics",
            "field": "statistical_adjustments",
            "type": "string",
            "source": "derived",
            "description": "Pipe-separated record of requested statistical adjustments applied at phase level.",
        },
        {
            "sheet": "repair_log",
            "field": "repair_notes",
            "type": "string",
            "source": "derived",
            "description": "Human-readable summary of the repairs applied to the session.",
        },
        {
            "sheet": "repair_log",
            "field": "external_straight_use_count_assumed",
            "type": "number",
            "source": "derived",
            "description": "For External sessions only: assumed straight-use count used in the session-level recalibration.",
        },
        {
            "sheet": "session_summary_raw",
            "field": "quality_flags",
            "type": "string",
            "source": "raw",
            "description": "Original raw-session quality flags before any repair.",
        },
        {
            "sheet": "phase_metrics_raw",
            "field": "raw_phase_id",
            "type": "string",
            "source": "raw",
            "description": "Original phase ID emitted by the exported JSON summary.",
        },
        {
            "sheet": "subjective_template",
            "field": "interruption_score",
            "type": "manual",
            "source": "paper / study design",
            "description": "Blank column for participant-rated interruption.",
        },
        {
            "sheet": "subjective_template",
            "field": "payload_quality",
            "type": "manual",
            "source": "paper / study design",
            "description": "Blank column for perceived suggestion payload quality if you decide to code it.",
        },
        {
            "sheet": "artifact_ratings_template",
            "field": "visual_organization",
            "type": "manual",
            "source": "paper / study design",
            "description": "Blank column for blind artifact rating.",
        },
        {
            "sheet": "artifact_ratings_template",
            "field": "depth_of_understanding",
            "type": "manual",
            "source": "paper / study design",
            "description": "Blank column for blind artifact rating.",
        },
        {
            "sheet": "artifact_ratings_template",
            "field": "breadth_coverage",
            "type": "manual",
            "source": "paper / study design",
            "description": "Blank column for blind artifact rating.",
        },
        {
            "sheet": "artifact_ratings_template",
            "field": "continuity_coherence",
            "type": "manual",
            "source": "paper / study design",
            "description": "Blank column for blind artifact rating.",
        },
        {
            "sheet": "artifact_ratings_template",
            "field": "later_review_usefulness",
            "type": "manual",
            "source": "paper / study design",
            "description": "Blank column for blind artifact rating.",
        },
    ]
    return pd.DataFrame(rows)


def make_review_sheet(session_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for row in session_df.to_dict(orient="records"):
        issues: list[str] = []
        repair_actions: list[str] = []
        notes: list[str] = []

        if row["quality_flags"] != "ok":
            if "missing_ended_at" in row["quality_flags"]:
                issues.append("missing_end_marker")
                repair_actions.append("Use best_effort_end_at_utc / best_effort_duration_seconds.")
                notes.append("Official end marker is missing.")
            if "has_failed_requests" in row["quality_flags"]:
                issues.append("failed_or_pending_requests")
                repair_actions.append("Prefer completed_round_count for efficiency/cost analyses.")
                notes.append("At least one AI round failed or remained pending at export.")
            if "incomplete_phase_coverage" in row["quality_flags"]:
                issues.append("phase_logging_incomplete")
                repair_actions.append("Do not trust logged phase IDs; only time-window imputation is possible.")
                notes.append("Not all three expected phase IDs were logged.")

        if row["external_ai_contaminated"]:
            issues.append("external_condition_contaminated")
            repair_actions.append("Exclude from pure External AI-behavior comparisons.")
            notes.append("External condition contains internal AI invocation(s).")
            if (row["accept_count"] or 0) == 0:
                notes.append("No internal AI suggestion was accepted, so the final artifact is likely unaffected.")

        if row["phase_input_artifacts_present"]:
            notes.append("Raw phase logs contain typing artifacts such as partial phase strings.")

        if (row["current_shape_count"] or 0) <= 2:
            notes.append("Final canvas is very sparse; inspect manually before artifact-quality scoring.")

        if not issues:
            issue_group = "clean"
            repairability = "none_needed"
            usable_session_level = "yes"
            usable_phase_level = "yes"
            usable_condition_comparison = "yes"
            usable_artifact_rating = "yes"
        else:
            issue_group = "|".join(issues)
            repairability = "repairable" if "phase_logging_incomplete" not in issues and "external_condition_contaminated" not in issues else "partial_only"
            usable_session_level = "yes_with_note"
            usable_phase_level = "impute_only" if "phase_logging_incomplete" in issues else "yes_with_note"
            usable_condition_comparison = "artifact_only" if "external_condition_contaminated" in issues else "yes_with_note"
            usable_artifact_rating = "yes_with_note"

        rows.append(
            {
                "session_id": row["session_id"],
                "topic": row["topic"],
                "condition": row["condition"],
                "source_file": row["source_file"],
                "issue_group": issue_group,
                "repairability": repairability,
                "usable_session_level": usable_session_level,
                "usable_phase_level": usable_phase_level,
                "usable_condition_comparison": usable_condition_comparison,
                "usable_artifact_rating": usable_artifact_rating,
                "best_effort_end_at_utc": row["best_effort_end_at_utc"],
                "best_effort_duration_seconds": row["best_effort_duration_seconds"],
                "repair_action": " ".join(repair_actions) if repair_actions else "No repair needed.",
                "review_notes": " ".join(notes) if notes else "Clean log.",
            }
        )

    review_df = pd.DataFrame(rows)
    if not review_df.empty:
        review_df["condition_order"] = review_df["condition"].map(CONDITION_ORDER)
        review_df = review_df.sort_values(
            ["topic", "condition_order", "session_id"],
            kind="stable",
        ).drop(columns=["condition_order"])
    return review_df


def apply_sheet_formatting(writer: pd.ExcelWriter) -> None:
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column_index, column_cells in enumerate(worksheet.iter_cols(1, worksheet.max_column), start=1):
            values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
            width = min(max(len(value) for value in values) + 2, 60)
            worksheet.column_dimensions[get_column_letter(column_index)].width = max(width, 12)


def export_workbook(
    output_path: Path,
    raw_session_df: pd.DataFrame,
    raw_phase_df: pd.DataFrame,
    final_session_df: pd.DataFrame,
    final_phase_df: pd.DataFrame,
    repair_df: pd.DataFrame,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    subjective_df = make_subjective_template(final_session_df)
    artifact_df = make_artifact_rating_template(final_session_df)
    dictionary_df = make_data_dictionary()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        final_session_df.to_excel(writer, sheet_name="session_summary", index=False)
        final_phase_df.to_excel(writer, sheet_name="phase_metrics", index=False)
        repair_df.to_excel(writer, sheet_name="repair_log", index=False)
        raw_session_df.to_excel(writer, sheet_name="session_summary_raw", index=False)
        raw_phase_df.to_excel(writer, sheet_name="phase_metrics_raw", index=False)
        subjective_df.to_excel(writer, sheet_name="subjective_template", index=False)
        artifact_df.to_excel(writer, sheet_name="artifact_ratings_template", index=False)
        dictionary_df.to_excel(writer, sheet_name="data_dictionary", index=False)
        apply_sheet_formatting(writer)


def main() -> None:
    args = parse_args()
    raw_session_df, raw_phase_df, final_session_df, final_phase_df, repair_df = make_session_rows(args.input_dir)
    export_workbook(
        args.output,
        raw_session_df=raw_session_df,
        raw_phase_df=raw_phase_df,
        final_session_df=final_session_df,
        final_phase_df=final_phase_df,
        repair_df=repair_df,
    )
    print(
        f"Wrote {len(final_session_df)} repaired sessions and {len(final_phase_df)} repaired phase rows to {args.output}"
    )


if __name__ == "__main__":
    main()
