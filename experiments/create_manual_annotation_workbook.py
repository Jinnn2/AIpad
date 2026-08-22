from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
EXPERIMENT_XLSX = ROOT / "results" / "experiment_important_variables_final.xlsx"
OUTPUT_XLSX = ROOT / "results" / "manual_annotation_workbook.xlsx"


def apply_sheet_formatting(writer: pd.ExcelWriter) -> None:
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column_index, column_cells in enumerate(worksheet.iter_cols(1, worksheet.max_column), start=1):
            values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
            width = min(max(len(value) for value in values) + 2, 42)
            worksheet.column_dimensions[get_column_letter(column_index)].width = max(width, 12)


def build_artifact_rating_sheet(session_df: pd.DataFrame) -> pd.DataFrame:
    raters = ["R1", "R2", "R3"]
    rows: list[dict[str, object]] = []
    for _, row in session_df.iterrows():
        for rater_id in raters:
            rows.append(
                {
                    "rater_id": rater_id,
                    "session_id": row["session_id"],
                    "participant_id": "",
                    "topic": row["topic"],
                    "condition": row["condition"],
                    "source_file": row["source_file"],
                    "visual_organization": "",
                    "depth_of_understanding": "",
                    "breadth_coverage": "",
                    "continuity_coherence": "",
                    "later_review_usefulness": "",
                    "artifact_notes": "",
                }
            )
    return pd.DataFrame(rows)


def build_interruption_sheet(session_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for _, row in session_df.iterrows():
        rows.append(
            {
                "participant_id": "",
                "session_id": row["session_id"],
                "topic": row["topic"],
                "condition": row["condition"],
                "source_file": row["source_file"],
                "interruption_score": "",
                "interruption_notes": "",
            }
        )
    return pd.DataFrame(rows)


def build_instructions_sheet() -> pd.DataFrame:
    instructions = [
        {
            "sheet": "artifact_ratings",
            "instruction": "Use the five rating columns for 10-point scores. One row = one rater x one session.",
        },
        {
            "sheet": "artifact_ratings",
            "instruction": "Suggested scale: 1 = very poor, 10 = excellent.",
        },
        {
            "sheet": "artifact_ratings",
            "instruction": "Keep `rater_id` as R1 / R2 / R3 unless you need a different naming scheme.",
        },
        {
            "sheet": "interruption_survey",
            "instruction": "One row = one session's participant survey response.",
        },
        {
            "sheet": "interruption_survey",
            "instruction": "Fill `participant_id` manually if you have a participant-session mapping.",
        },
        {
            "sheet": "interruption_survey",
            "instruction": "Use `interruption_score` for the main numeric interruption item and `interruption_notes` for any comments or coding notes.",
        },
    ]
    return pd.DataFrame(instructions)


def main() -> None:
    session_df = pd.read_excel(EXPERIMENT_XLSX, sheet_name="session_summary")
    session_df = session_df[["session_id", "topic", "condition", "source_file"]].copy()

    artifact_df = build_artifact_rating_sheet(session_df)
    interruption_df = build_interruption_sheet(session_df)
    instructions_df = build_instructions_sheet()

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        instructions_df.to_excel(writer, sheet_name="instructions", index=False)
        artifact_df.to_excel(writer, sheet_name="artifact_ratings", index=False)
        interruption_df.to_excel(writer, sheet_name="interruption_survey", index=False)
        apply_sheet_formatting(writer)

    print(f"Wrote manual annotation workbook to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
